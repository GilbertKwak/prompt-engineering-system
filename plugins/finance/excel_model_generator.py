#!/usr/bin/env python3
"""
excel_model_generator.py  v1.0
PE-7 P2-1: AstraChips LP Fund 재무 모델 자동 Excel 생성
입력: cowork_finance_runner.py 쮜력 JSON
산출물: 7시트 XLSX (Summary / Scenario x4 / Monte Carlo / Risk Matrix)
작성: 2026-04-26 | Gilbert Kwak
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────
# 컴팟 2026 디자인 코드
# ──────────────────────────────────────────────────────────
COLOR = {
    "navy":        "1B2A4A",
    "blue":        "2563EB",
    "blue_light":  "DBEAFE",
    "teal":        "0D9488",
    "teal_light":  "CCFBF1",
    "amber":       "D97706",
    "amber_light": "FEF3C7",
    "red":         "DC2626",
    "red_light":   "FEE2E2",
    "green":       "16A34A",
    "green_light": "DCFCE7",
    "gray_dark":   "374151",
    "gray_mid":    "6B7280",
    "gray_light":  "F9FAFB",
    "white":       "FFFFFF",
    "border":      "D1D5DB",
}

TYPE_COLOR = {
    "type_a": (COLOR["blue"],       COLOR["blue_light"]),
    "type_b": (COLOR["teal"],       COLOR["teal_light"]),
    "type_c": (COLOR["amber"],      COLOR["amber_light"]),
}

SCEN_ORDER = ["balanced", "conservative", "aggressive", "geo_hedge"]
SCEN_LABEL = {
    "balanced":     "Balanced (Base)",
    "conservative": "Conservative",
    "aggressive":   "Aggressive",
    "geo_hedge":    "Geo-Hedge",
}
SCEN_COLOR = {
    "balanced":     COLOR["blue"],
    "conservative": COLOR["teal"],
    "aggressive":   COLOR["red"],
    "geo_hedge":    COLOR["amber"],
}


# ──────────────────────────────────────────────────────────
# 스타일 헬퍼
# ──────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")


def _border_thin() -> Border:
    s = Side(style="thin", color=COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _border_medium() -> Border:
    s = Side(style="medium", color=COLOR["gray_dark"])
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_header_cell(ws, row, col, value, bg=None, font=None, border=None,
                       align=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:     cell.fill = bg
    if font:   cell.font = font
    if border: cell.border = border
    if align:  cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    return cell


def _section_title(ws, row, col, text, span, bg_hex=None):
    """A~B+span 열 머지 상단 제목 바."""
    bg = _fill(bg_hex or COLOR["navy"])
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = bg
    cell.font = _font(bold=True, size=12, color=COLOR["white"])
    cell.alignment = _align("left")
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    return row + 1


def _table_header_row(ws, row, headers: list, bg_hex: str,
                      start_col: int = 1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = _fill(bg_hex)
        c.font = _font(bold=True, size=10, color=COLOR["white"])
        c.alignment = _align("center")
        c.border = _border_thin()
    return row + 1


def _data_row(ws, row, values: list, bg_hex: str = None,
              start_col: int = 1, num_fmts: list = None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.fill = _fill(bg_hex or COLOR["white"])
        c.font = _font(size=10)
        c.alignment = _align("center")
        c.border = _border_thin()
        if num_fmts and i < len(num_fmts) and num_fmts[i]:
            c.number_format = num_fmts[i]
    return row + 1


def _col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ──────────────────────────────────────────────────────────
# Sheet 1: Summary Dashboard
# ──────────────────────────────────────────────────────────
def build_sheet_summary(wb, data: dict):
    ws = wb.active
    ws.title = "1. Summary"
    ws.sheet_view.showGridLines = False

    meta = data["meta"]
    summary = data.get("portfolio_summary", {})
    scenarios = data["scenarios"]

    # ─ 타이틀 대내장 ─
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "AstraChips LP Fund — Portfolio Finance Model"
    title_cell.fill = _fill(COLOR["navy"])
    title_cell.font = _font(bold=True, size=16, color=COLOR["white"])
    title_cell.alignment = _align("center")
    _row_height(ws, 1, 36)

    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    sub.value = (
        f"Generated: {meta['generated_at'][:10]}  |  "
        f"Fund Size: ${meta['fund_size_M']:,.0f}M  |  "
        f"Discount Rate: {meta['discount_rate']*100:.0f}%  |  "
        f"Horizon: {meta['projection_years']}Y  |  PE {meta['pe_version']}"
    )
    sub.fill = _fill(COLOR["gray_dark"])
    sub.font = _font(size=10, color=COLOR["white"], italic=True)
    sub.alignment = _align("center")
    _row_height(ws, 2, 20)

    row = 4

    # ─ KPI 카드 (3개: Fund Size / Expected MOIC / Base Scenario) ─
    kpi_data = [
        ("Total Fund Size",     f"${meta['fund_size_M']:,.0f}M",   COLOR["blue"]),
        ("Expected MOIC",       f"{summary.get('expected_moic', 'N/A')}x",  COLOR["teal"]),
        ("Base Scenario",       summary.get('base_scenario', 'balanced').title(), COLOR["amber"]),
    ]
    kpi_cols = [1, 4, 7]
    for (label, value, color), sc in zip(kpi_data, kpi_cols):
        # 메인 값
        ws.merge_cells(start_row=row, start_column=sc,
                       end_row=row, end_column=sc + 2)
        c = ws.cell(row=row, column=sc, value=value)
        c.fill = _fill(color)
        c.font = _font(bold=True, size=20, color=COLOR["white"])
        c.alignment = _align("center")
        # 레이블
        ws.merge_cells(start_row=row+1, start_column=sc,
                       end_row=row+1, end_column=sc + 2)
        lc = ws.cell(row=row+1, column=sc, value=label)
        lc.fill = _fill(COLOR["gray_light"])
        lc.font = _font(size=9, color=COLOR["gray_mid"])
        lc.alignment = _align("center")
        _row_height(ws, row, 32)
        _row_height(ws, row+1, 18)

    row += 3

    # ─ 시나리오 MOIC 비교 테이블 ─
    row = _section_title(ws, row, 1, "📊 Portfolio MOIC by Scenario", 9)
    headers = ["Scenario", "Allocation (A/B/C)", "Probability",
               "Portfolio MOIC", "Type A MOIC", "Type B MOIC", "Type C MOIC",
               "Total Invested ($M)", "Risk Score"]
    row = _table_header_row(ws, row, headers, COLOR["gray_dark"])

    alt = False
    for sk in SCEN_ORDER:
        if sk not in scenarios:
            continue
        sd = scenarios[sk]
        alloc = sd["types"]
        a = alloc.get("type_a", {})
        b = alloc.get("type_b", {})
        c = alloc.get("type_c", {})
        alloc_str = (
            f"{int(a.get('weight',0)*100)}/"
            f"{int(b.get('weight',0)*100)}/"
            f"{int(c.get('weight',0)*100)}"
        )
        bg = COLOR["gray_light"] if alt else COLOR["white"]
        vals = [
            SCEN_LABEL.get(sk, sk),
            alloc_str,
            f"{sd.get('probability',0)*100:.0f}%",
            f"{sd.get('portfolio_moic', 0):.2f}x",
            f"{a.get('moic', 0):.2f}x",
            f"{b.get('moic', 0):.2f}x",
            f"{c.get('moic', 0):.2f}x",
            f"${sd.get('total_invested_M', 0):,.0f}M",
            f"{sd.get('risk_score', 0):.3f}",
        ]
        row = _data_row(ws, row, vals, bg)
        alt = not alt

    row += 1

    # ─ Type 별 기본 시나리오 상세 ─
    base_scen = scenarios.get("balanced", {})
    row = _section_title(ws, row, 1,
                         "🎯 Base Scenario (Balanced) — Type Breakdown", 9)
    headers2 = ["Type", "Strategy", "Weight", "Investment ($M)",
                "IRR Base", "IRR Bull", "IRR Bear",
                "NPV ($M)", "MOIC Target"]
    row = _table_header_row(ws, row, headers2, COLOR["blue"])

    for tk, (bg_main, bg_light) in TYPE_COLOR.items():
        td = base_scen.get("types", {}).get(tk, {})
        vals = [
            tk.upper().replace("_", " "),
            td.get("label", ""),
            f"{td.get('weight',0)*100:.0f}%",
            f"${td.get('investment_M',0):,.0f}M",
            f"{td.get('irr_base_pct',0):.1f}%",
            f"{td.get('irr_bull_pct',0):.1f}%",
            f"{td.get('irr_bear_pct',0):.1f}%",
            f"${td.get('npv_M',0):,.1f}M",
            f"{td.get('moic',0):.2f}x",
        ]
        row = _data_row(ws, row, vals, bg_light)

    row += 1

    # ─ 열 너비 ─
    _col_widths(ws, {
        "A": 20, "B": 30, "C": 14, "D": 18, "E": 14,
        "F": 14, "G": 14, "H": 18, "I": 14, "J": 4
    })

    # ─ Portfolio MOIC 막대 차트 ─
    chart = BarChart()
    chart.type = "col"
    chart.title = "Portfolio MOIC by Scenario"
    chart.y_axis.title = "MOIC (x)"
    chart.x_axis.title = "Scenario"
    chart.style = 10
    chart.width = 22
    chart.height = 12

    # 데이터 범위 동적 참조 (MOIC 열)
    moic_row_start = 10  # 시나리오 테이블 데이터 시작 예시
    n_scens = sum(1 for sk in SCEN_ORDER if sk in scenarios)
    data_ref = Reference(ws, min_col=4, min_row=moic_row_start,
                         max_row=moic_row_start + n_scens - 1)
    cats_ref = Reference(ws, min_col=1, min_row=moic_row_start,
                         max_row=moic_row_start + n_scens - 1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.series[0].title.v = "Portfolio MOIC"

    # 막대 색상 개별 지정
    scen_hex = [SCEN_COLOR.get(sk) for sk in SCEN_ORDER if sk in scenarios]
    for idx, hex_c in enumerate(scen_hex):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = hex_c
        chart.series[0].dPt.append(pt)

    anchor_row = row + 1
    ws.add_chart(chart, f"A{anchor_row}")

    print("  ✅ Sheet 1: Summary Dashboard")


# ──────────────────────────────────────────────────────────
# Sheet 2~5: 시나리오별 현금흐름 + IRR/NPV 상세
# ──────────────────────────────────────────────────────────
def build_sheet_scenario(wb, scen_key: str, scen_idx: int, data: dict):
    """시나리오 1개 = 1시트."""
    scen = data["scenarios"].get(scen_key, {})
    meta = data["meta"]
    label = SCEN_LABEL.get(scen_key, scen_key)
    color = SCEN_COLOR.get(scen_key, COLOR["navy"])

    ws = wb.create_sheet(title=f"{scen_idx}. {label[:12]}")
    ws.sheet_view.showGridLines = False

    # ─ 타이틀 ─
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value = f"Scenario: {label}  |  Portfolio MOIC: {scen.get('portfolio_moic', 0):.2f}x  |  Risk Score: {scen.get('risk_score', 0):.3f}"
    tc.fill = _fill(color)
    tc.font = _font(bold=True, size=13, color=COLOR["white"])
    tc.alignment = _align("center")
    _row_height(ws, 1, 30)

    row = 3

    # ─ 타입별 파라미터 테이블 ─
    row = _section_title(ws, row, 1, f"📋 Type Allocation & Return Metrics", 11, color)
    headers = ["Type", "Label", "Weight", "Invest ($M)",
               "IRR Base%", "IRR Bull%", "IRR Bear%",
               "IRR Calc%", "NPV ($M)", "Proceeds ($M)", "MOIC"]
    row = _table_header_row(ws, row, headers, color)

    for tk, (bg_main, bg_light) in TYPE_COLOR.items():
        td = scen.get("types", {}).get(tk, {})
        vals = [
            tk.upper().replace("_", " "),
            td.get("label", ""),
            f"{td.get('weight', 0)*100:.0f}%",
            f"${td.get('investment_M', 0):,.0f}M",
            f"{td.get('irr_base_pct', 0):.1f}%",
            f"{td.get('irr_bull_pct', 0):.1f}%",
            f"{td.get('irr_bear_pct', 0):.1f}%",
            f"{td.get('irr_calc_pct', 0):.1f}%",
            f"${td.get('npv_M', 0):,.1f}M",
            f"${td.get('total_proceeds_M', 0):,.1f}M",
            f"{td.get('moic', 0):.2f}x",
        ]
        row = _data_row(ws, row, vals, bg_light)

    # ─ 토탈 행 ─
    total_vals = [
        "TOTAL", "", "100%",
        f"${scen.get('total_invested_M', 0):,.0f}M",
        "", "", "", "",
        "",
        f"${scen.get('total_proceeds_M', 0):,.1f}M",
        f"{scen.get('portfolio_moic', 0):.2f}x",
    ]
    r = row
    for i, v in enumerate(total_vals):
        c = ws.cell(row=r, column=1+i, value=v)
        c.fill = _fill(color)
        c.font = _font(bold=True, size=10, color=COLOR["white"])
        c.alignment = _align("center")
        c.border = _border_thin()
    row += 2

    # ─ 타입별 J-Curve 현금흐름 ─
    row = _section_title(ws, row, 1, "📈 J-Curve Cash Flows by Type ($M)", 11, color)

    # 헤더: Year 열 동적 생성
    sample_type = next(iter(scen.get("types", {}).values()), {})
    years = sample_type.get("cash_flows", {}).get("years", [])
    cf_headers = ["Type"] + [str(y) for y in years] + ["Total Net"]
    row = _table_header_row(ws, row, cf_headers, color)

    type_cf_rows = {}  # 나중에 차트 용
    cf_data_start_row = row

    for tk, (bg_main, bg_light) in TYPE_COLOR.items():
        td = scen.get("types", {}).get(tk, {})
        cfs = td.get("cash_flows", {}).get("cash_flows_M", [])
        net = sum(cfs)
        vals = [tk.upper().replace("_", " ")] + [round(v, 2) for v in cfs] + [round(net, 2)]
        type_cf_rows[tk] = row
        row = _data_row(ws, row, vals, bg_light)

    row += 1

    # J-Curve 라인 차트
    chart = LineChart()
    chart.title = f"{label} — J-Curve Cash Flows"
    chart.y_axis.title = "Cash Flow ($M)"
    chart.x_axis.title = "Year"
    chart.style = 10
    chart.width = 24
    chart.height = 12

    n_years = len(years)
    for tk_idx, tk in enumerate(["type_a", "type_b", "type_c"]):
        cf_row = type_cf_rows.get(tk)
        if cf_row is None:
            continue
        # 칼럼 2~(2+n_years)
        ref = Reference(ws, min_col=2, max_col=1+n_years,
                        min_row=cf_row, max_row=cf_row)
        chart.add_data(ref, titles_from_data=False)
        chart.series[tk_idx].title.v = tk.upper().replace("_", " ")
        chart.series[tk_idx].graphicalProperties.line.solidFill = \
            list(TYPE_COLOR.values())[tk_idx][0]
        chart.series[tk_idx].graphicalProperties.line.width = 25000

    # X축 레이블
    cats = Reference(ws, min_col=2, max_col=1+n_years,
                     min_row=3, max_row=3)  # 로우 3은 헤더 행
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{row}")

    # ─ 열 너비 ─
    _col_widths(ws, {
        "A": 18, "B": 30, "C": 12, "D": 14, "E": 12,
        "F": 12, "G": 12, "H": 12, "I": 14, "J": 14,
        "K": 12, "L": 4,
    })

    print(f"  ✅ Sheet {scen_idx}: {label}")


# ──────────────────────────────────────────────────────────
# Sheet 6: Monte Carlo 분위수
# ──────────────────────────────────────────────────────────
def build_sheet_monte_carlo(wb, data: dict):
    ws = wb.create_sheet(title="6. Monte Carlo")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value = "Monte Carlo Simulation — IRR Percentile Distribution (10,000 iterations)"
    tc.fill = _fill(COLOR["navy"])
    tc.font = _font(bold=True, size=13, color=COLOR["white"])
    tc.alignment = _align("center")
    _row_height(ws, 1, 30)

    row = 3
    scenarios = data["scenarios"]

    # ─ 시나리오 × Type 전체 MC 테이블 ─
    row = _section_title(ws, row, 1,
                         "🎲 IRR Percentile Table by Scenario & Type", 11)
    headers = [
        "Scenario", "Type",
        "P10 IRR%", "P25 IRR%", "P50 IRR%", "P75 IRR%", "P90 IRR%",
        "Mean IRR%", "Std Dev%",
        "P(IRR>8%)", "P(IRR>20%)"
    ]
    row = _table_header_row(ws, row, headers, COLOR["gray_dark"])

    mc_table_start = row  # 차트 데이터 시작
    alt = False
    for sk in SCEN_ORDER:
        if sk not in scenarios:
            continue
        sd = scenarios[sk]
        scen_label = SCEN_LABEL.get(sk, sk)
        scen_color = SCEN_COLOR.get(sk, COLOR["navy"])

        for tk, (bg_main, bg_light) in TYPE_COLOR.items():
            td = sd.get("types", {}).get(tk, {})
            mc = td.get("monte_carlo", {})
            bg = COLOR["gray_light"] if alt else COLOR["white"]
            vals = [
                scen_label,
                tk.upper().replace("_", " "),
                f"{mc.get('p10', 0)*100:.1f}%",
                f"{mc.get('p25', 0)*100:.1f}%",
                f"{mc.get('p50', 0)*100:.1f}%",
                f"{mc.get('p75', 0)*100:.1f}%",
                f"{mc.get('p90', 0)*100:.1f}%",
                f"{mc.get('mean', 0)*100:.1f}%",
                f"{mc.get('std', 0)*100:.1f}%",
                f"{mc.get('prob_above_hurdle_8pct', 0)*100:.0f}%",
                f"{mc.get('prob_above_20pct', 0)*100:.0f}%",
            ]
            row = _data_row(ws, row, vals, bg)
            alt = not alt

    row += 1

    # ─ Base 시나리오 IRR 대표값 요약 ─
    row = _section_title(ws, row, 1,
                         "⭐ Base Scenario (Balanced) MC Summary", 11,
                         COLOR["blue"])
    base = scenarios.get("balanced", {})
    sum_headers = ["Type", "Label",
                   "Bear (P10)", "P25", "Median (P50)",
                   "P75", "Bull (P90)",
                   "P(>Hurdle 8%)", "P(>20%)"]
    row = _table_header_row(ws, row, sum_headers, COLOR["blue"])

    base_mc_start = row
    for tk, (bg_main, bg_light) in TYPE_COLOR.items():
        td = base.get("types", {}).get(tk, {})
        mc = td.get("monte_carlo", {})
        vals = [
            tk.upper().replace("_", " "),
            td.get("label", ""),
            f"{mc.get('p10', 0)*100:.1f}%",
            f"{mc.get('p25', 0)*100:.1f}%",
            f"{mc.get('p50', 0)*100:.1f}%",
            f"{mc.get('p75', 0)*100:.1f}%",
            f"{mc.get('p90', 0)*100:.1f}%",
            f"{mc.get('prob_above_hurdle_8pct', 0)*100:.0f}%",
            f"{mc.get('prob_above_20pct', 0)*100:.0f}%",
        ]
        row = _data_row(ws, row, vals, bg_light)

    row += 1

    # P10~P90 범위 차트 (Bar: Base 시나리오 타입별)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "IRR Percentile Range — Base Scenario by Type"
    chart2.y_axis.title = "IRR (%)"
    chart2.style = 10
    chart2.width = 24
    chart2.height = 13

    # P10, P50, P90 3개 시리즈
    for col_offset, pct_label in [(3, "P10"), (5, "P50"), (7, "P90")]:
        ref = Reference(ws, min_col=col_offset, max_col=col_offset,
                        min_row=base_mc_start, max_row=base_mc_start + 2)
        chart2.add_data(ref, titles_from_data=False)
        series = chart2.series[-1]
        series.title.v = pct_label

    cat_ref = Reference(ws, min_col=1, min_row=base_mc_start,
                        max_row=base_mc_start + 2)
    chart2.set_categories(cat_ref)
    ws.add_chart(chart2, f"A{row}")

    _col_widths(ws, {
        "A": 20, "B": 32, "C": 12, "D": 12, "E": 12,
        "F": 12, "G": 12, "H": 14, "I": 12, "J": 14,
        "K": 12, "L": 4
    })

    print("  ✅ Sheet 6: Monte Carlo")


# ──────────────────────────────────────────────────────────
# Sheet 7: Risk Matrix
# ──────────────────────────────────────────────────────────
def build_sheet_risk_matrix(wb, data: dict):
    ws = wb.create_sheet(title="7. Risk Matrix")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    tc = ws["A1"]
    tc.value = "AstraChips LP Fund — Risk Factor Matrix"
    tc.fill = _fill(COLOR["navy"])
    tc.font = _font(bold=True, size=13, color=COLOR["white"])
    tc.alignment = _align("center")
    _row_height(ws, 1, 30)

    row = 3

    # ─ 리스크 요인 테이블 ─
    RISK_DATA = [
        # (category, factor, probability, impact, type_exposure, mitigation)
        ("🌐 Geopolitical", "Korea Peninsula Escalation",    "15%", "High",   "A",   "JV structure with exit clause"),
        ("🌐 Geopolitical", "Taiwan Strait Blockade",         "20%", "Critical","B",  "Offtake + dual-source contract"),
        ("🌐 Geopolitical", "US/EU Policy Reversal",          "10%", "Medium", "C",   "CHIPS Act lock-in provisions"),
        ("💹 Market",       "HBM Cycle Downturn",             "25%", "High",   "A+B", "Hedged via Type C"),
        ("💹 Market",       "Glass Substrate Yield Delay",    "15%", "High",   "A",   "Samsung backup + AGC ramp"),
        ("💹 Market",       "Competitor Technology Leap",     "12%", "Medium", "B",   "HBM5 roadmap option"),
        ("⚙️ Execution",   "JV Negotiation Failure",         "10%", "High",   "A",   "Pre-agreed term sheet"),
        ("⚙️ Execution",   "Offtake Contract Renegotiation", "8%",  "Medium", "B",   "5Y take-or-pay clause"),
        ("⚙️ Execution",   "Regulatory Approval Delay",      "12%", "Medium", "C",   "CFIUS pre-clearance"),
    ]
    IMPACT_COLOR_MAP = {
        "Critical": COLOR["red"],
        "High":     COLOR["amber"],
        "Medium":   COLOR["teal"],
        "Low":      COLOR["green"],
    }

    row = _section_title(ws, row, 1, "🚨 Risk Factor Register", 9)
    headers = ["Category", "Risk Factor", "Probability",
               "Impact", "Type Exposure", "Mitigation Strategy"]
    row = _table_header_row(ws, row, headers, COLOR["gray_dark"])

    alt = False
    for rdata in RISK_DATA:
        category, factor, prob, impact, exposure, mitigation = rdata
        bg = COLOR["gray_light"] if alt else COLOR["white"]

        for i, v in enumerate([category, factor, prob, impact, exposure, mitigation]):
            c = ws.cell(row=row, column=1+i, value=v)
            c.border = _border_thin()
            c.alignment = _align("center" if i != 5 else "left", wrap=True)
            c.font = _font(size=10)
            # Impact 열은 색상차로
            if i == 3:
                c.fill = _fill(IMPACT_COLOR_MAP.get(impact, COLOR["gray_light"]))
                c.font = _font(bold=True, size=10, color=COLOR["white"])
            else:
                c.fill = _fill(bg)
        row += 1
        alt = not alt

    row += 1

    # ─ 시나리오별 리스크 점수 비교 ─
    row = _section_title(ws, row, 1,
                         "📊 Scenario Risk Score Comparison", 5,
                         COLOR["blue"])
    rsk_headers = ["Scenario", "Allocation (A/B/C)",
                   "Risk Score", "Probability", "Recommended"]
    row = _table_header_row(ws, row, rsk_headers, COLOR["blue"])

    scen_risk_start = row
    scenarios = data["scenarios"]
    for sk in SCEN_ORDER:
        if sk not in scenarios:
            continue
        sd = scenarios[sk]
        alloc = sd["types"]
        a = alloc.get("type_a", {})
        b = alloc.get("type_b", {})
        c_ = alloc.get("type_c", {})
        alloc_str = (
            f"{int(a.get('weight',0)*100)}/"
            f"{int(b.get('weight',0)*100)}/"
            f"{int(c_.get('weight',0)*100)}"
        )
        is_base = sk == "balanced"
        rec_val = "⭐ RECOMMENDED" if is_base else ""

        risk_score = sd.get("risk_score", 0)
        # 리스크 점수에 따른 색상
        if risk_score < 0.15:
            rsk_color = COLOR["green"]
        elif risk_score < 0.18:
            rsk_color = COLOR["amber"]
        else:
            rsk_color = COLOR["red"]

        for i, v in enumerate([
            SCEN_LABEL.get(sk, sk),
            alloc_str,
            f"{risk_score:.3f}",
            f"{sd.get('probability',0)*100:.0f}%",
            rec_val,
        ]):
            c = ws.cell(row=row, column=1+i, value=v)
            c.border = _border_thin()
            c.alignment = _align("center")
            if i == 2:
                c.fill = _fill(rsk_color)
                c.font = _font(bold=True, size=10, color=COLOR["white"])
            elif is_base:
                c.fill = _fill(COLOR["blue_light"])
                c.font = _font(bold=True, size=10)
            else:
                c.fill = _fill(COLOR["white"])
                c.font = _font(size=10)
        row += 1

    row += 1

    # 리스크 점수 막대 차트
    chart3 = BarChart()
    chart3.type = "bar"   # 가로막대
    chart3.title = "Risk Score by Scenario"
    chart3.x_axis.title = "Risk Score"
    chart3.style = 10
    chart3.width = 18
    chart3.height = 10

    n_s = sum(1 for sk in SCEN_ORDER if sk in scenarios)
    rsk_ref = Reference(ws, min_col=3, min_row=scen_risk_start,
                        max_row=scen_risk_start + n_s - 1)
    cat_ref2 = Reference(ws, min_col=1, min_row=scen_risk_start,
                         max_row=scen_risk_start + n_s - 1)
    chart3.add_data(rsk_ref, titles_from_data=False)
    chart3.set_categories(cat_ref2)
    chart3.series[0].title.v = "Risk Score"

    ws.add_chart(chart3, f"A{row}")

    _col_widths(ws, {
        "A": 20, "B": 30, "C": 14,
        "D": 16, "E": 20, "F": 40,
    })

    print("  ✅ Sheet 7: Risk Matrix")


# ──────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Excel Model Generator v1.0 — AstraChips LP Fund"
    )
    parser.add_argument(
        "--input",
        default="reports/finance_output.json",
        help="cowork_finance_runner.py 쮜력 JSON 경로"
    )
    parser.add_argument(
        "--output",
        default="reports/AstraChips_Finance_Model.xlsx",
        help="출력 XLSX 경로"
    )
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"[ERROR] Input JSON not found: {in_path}")
        print("  먼저 cowork_finance_runner.py를 실행하세요.")
        sys.exit(1)

    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)

    print(f"\n📂 Finance data loaded: {in_path}")
    print(f"   Scenarios: {list(data['scenarios'].keys())}")
    print(f"   Fund Size: ${data['meta']['fund_size_M']:,.0f}M\n")
    print("📊 Building Excel model...")

    # 강제 실제 시나리오 키 정렬 (JSON 순서 문제 대비)
    scen_keys_ordered = [sk for sk in SCEN_ORDER
                         if sk in data["scenarios"]]

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    build_sheet_summary(wb, data)

    # Sheet 2~5: 시나리오 상세
    for idx, sk in enumerate(scen_keys_ordered, start=2):
        build_sheet_scenario(wb, sk, idx, data)

    # Sheet 6: Monte Carlo
    build_sheet_monte_carlo(wb, data)

    # Sheet 7: Risk Matrix
    build_sheet_risk_matrix(wb, data)

    # 출력
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"\n✅ Excel model saved: {out_path}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
