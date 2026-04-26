#!/usr/bin/env python3
"""
smr_power_model.py  v1.0
PE-7 P3: SMR 전력비 모델 — Excel Sheet 8~9 생성기
기존 excel_model_generator.py v1.0 (Sheet 1~7)에 Sheet 8~9 추가

사용법:
  python plugins/finance/smr_power_model.py
  python plugins/finance/smr_power_model.py --input reports/AstraChips_Finance_Model.xlsx
  python plugins/finance/smr_power_model.py --dry-run

작성: 2026-04-26 | Gilbert Kwak | PE-7 P3
"""

import argparse
import json
import math
import sys
from datetime import datetime
from pathlib import Path

import yaml

try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────────────────
# 공용 색상 / 스타일 (excel_model_generator.py 와 동일)
# ──────────────────────────────────────────────────────────
COLOR = {
    "navy":        "1B2A4A",
    "blue":        "2563EB",
    "blue_light":  "DBEAFE",
    "teal":        "0D9488",
    "teal_light":  "CCFBF1",
    "amber":       "D97706",
    "amber_light": "FEF3C7",
    "red":         "DC2626",
    "red_light":   "FEE2E2",
    "green":       "16A34A",
    "green_light": "DCFCE7",
    "purple":      "7C3AED",
    "purple_light":"EDE9FE",
    "gray_dark":   "374151",
    "gray_mid":    "6B7280",
    "gray_light":  "F9FAFB",
    "white":       "FFFFFF",
    "border":      "D1D5DB",
}


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _border_thin():
    s = Side(style="thin", color=COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def _row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def _section_title(ws, row, col, text, span, bg_hex=None):
    bg = _fill(bg_hex or COLOR["navy"])
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = bg
    cell.font = _font(bold=True, size=12, color=COLOR["white"])
    cell.alignment = _align("left")
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    return row + 1

def _table_header_row(ws, row, headers, bg_hex, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = _fill(bg_hex)
        c.font = _font(bold=True, size=10, color=COLOR["white"])
        c.alignment = _align("center")
        c.border = _border_thin()
    return row + 1

def _data_row(ws, row, values, bg_hex=None, start_col=1, bold=False):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.fill = _fill(bg_hex or COLOR["white"])
        c.font = _font(bold=bold, size=10)
        c.alignment = _align("center")
        c.border = _border_thin()
    return row + 1


# ──────────────────────────────────────────────────────────
# SMR 재무 계산 헬퍼
# ──────────────────────────────────────────────────────────
def calc_lcoe(smr: dict) -> float:
    """LCOE = (CAPEX annuity + OPEX + Fuel) / 연간 발전량"""
    capex_usd = smr["capex_b_usd"] * 1e9
    cap_mwe = smr["capacity_mwe"]
    af = smr["availability_factor"]
    life_y = smr["lifetime_years"]
    r = 0.07  # weighted cost of capital for SMR

    annual_mwh = cap_mwe * af * 8760  # MWh/year
    # CAPEX 연금 현가
    annuity_factor = r / (1 - (1 + r) ** (-life_y))
    capex_annual = capex_usd * annuity_factor
    capex_per_mwh = capex_annual / annual_mwh

    lcoe = capex_per_mwh + smr["opex_per_mwh"] + smr["fuel_per_mwh"]
    return round(lcoe, 2)


def calc_npv_savings(grid_price, smr_lcoe, smr_pct, annual_gwh, discount_rate, years):
    """그리드 대비 SMR 전력비 절감 NPV"""
    delta_per_mwh = grid_price - smr_lcoe  # $/MWh 절감액
    smr_gwh = annual_gwh * smr_pct
    annual_saving = delta_per_mwh * smr_gwh * 1e3  # $M (GWh→MWh)
    npv = sum(annual_saving / ((1 + discount_rate) ** t) for t in range(1, years + 1))
    return round(npv / 1e6, 1)  # $M


def calc_grid_escalated_cost(grid_price, escalation, annual_gwh, years):
    """그리드 전력비 10년 누적 비용 (에스컬레이션 포함)"""
    total = sum(
        grid_price * (1 + escalation) ** t * annual_gwh * 1e3
        for t in range(years)
    )
    return round(total / 1e6, 1)  # $M


# ──────────────────────────────────────────────────────────
# Sheet 8: SMR Power Cost Analysis
# ──────────────────────────────────────────────────────────
def build_sheet_smr_power(wb, cfg: dict):
    ws = wb.create_sheet(title="8. SMR Power Cost")
    ws.sheet_view.showGridLines = False

    meta = cfg["meta"]
    smr_types = cfg["smr_types"]
    dc_tiers = cfg["data_center_tiers"]
    scenarios = cfg["scenarios"]
    sens = cfg["sensitivity"]
    dr = meta["discount_rate"]
    proj_y = meta["projection_years"]

    # ─ 타이틀 ─
    ws.merge_cells("A1:N1")
    tc = ws["A1"]
    tc.value = "SMR Power Cost Analysis — AstraChips LP Fund | PE-7 P3"
    tc.fill = _fill(COLOR["navy"])
    tc.font = _font(bold=True, size=15, color=COLOR["white"])
    tc.alignment = _align("center")
    _row_height(ws, 1, 36)

    ws.merge_cells("A2:N2")
    sub = ws["A2"]
    sub.value = (
        f"Base Year: {meta['base_year']}  |  "
        f"Discount Rate: {dr*100:.0f}%  |  "
        f"Projection: {proj_y}Y  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    sub.fill = _fill(COLOR["gray_dark"])
    sub.font = _font(size=10, color=COLOR["white"], italic=True)
    sub.alignment = _align("center")
    _row_height(ws, 2, 18)

    row = 4

    # ─ [1] SMR 기술 비교 테이블 ─
    row = _section_title(ws, row, 1, "⚛️  SMR Technology Comparison", 12)
    headers1 = [
        "SMR Type", "Vendor", "Capacity\n(MWe)", "CAPEX\n($B)",
        "CAPEX/kWe\n($/kWe)", "OPEX\n($/MWh)", "Fuel\n($/MWh)",
        "LCOE\n($/MWh)", "Availability", "CO₂\n(kg/MWh)",
        "Status", "COD Target"
    ]
    row = _table_header_row(ws, row, headers1, COLOR["purple"])

    smr_keys = list(smr_types.keys())
    smr_lcoe_map = {}  # 이후 계산에 재사용
    for i, sk in enumerate(smr_keys):
        sm = smr_types[sk]
        lcoe = calc_lcoe(sm)
        smr_lcoe_map[sk] = lcoe
        bg = COLOR["purple_light"] if i % 2 == 0 else COLOR["white"]
        vals = [
            sm["name"], sm["vendor"],
            f"{sm['capacity_mwe']:,}",
            f"${sm['capex_b_usd']:.1f}B",
            f"${sm['capex_per_kwe']:,}",
            f"${sm['opex_per_mwh']:.1f}",
            f"${sm['fuel_per_mwh']:.1f}",
            f"${lcoe:.0f}",
            f"{sm['availability_factor']*100:.0f}%",
            f"{sm['co2_kg_per_mwh']:.1f}",
            sm["status"],
            sm["notes"].split(".")[0],
        ]
        row = _data_row(ws, row, vals, bg)

    row += 1

    # ─ [2] Tier × SMR 전력비 비교 테이블 ─
    row = _section_title(ws, row, 1,
        "💡 Grid vs SMR Power Cost by Facility Tier ($M over 10Y)", 12,
        COLOR["teal"])
    headers2 = [
        "Facility Tier", "Type Link",
        "Annual Power\n(GWh)", "Grid Price\n($/MWh)",
        "Grid 10Y Cost\n($M)",
        "NuScale LCOE", "NuScale NPV Saving",
        "RR SMR LCOE", "RR SMR NPV Saving",
        "Kairos LCOE", "Kairos NPV Saving",
        "Best SMR",
    ]
    row = _table_header_row(ws, row, headers2, COLOR["teal"])

    tier_smr_table_start = row
    alt = False
    for tk, td in dc_tiers.items():
        grid_10y = calc_grid_escalated_cost(
            td["grid_price_per_mwh"],
            td["grid_price_escalation"],
            td["annual_power_gwh"],
            proj_y,
        )
        savings = {}
        for sk, lcoe in smr_lcoe_map.items():
            npv_sav = calc_npv_savings(
                td["grid_price_per_mwh"],
                lcoe, 0.8,  # SMR Full 80%
                td["annual_power_gwh"],
                dr, proj_y,
            )
            savings[sk] = (lcoe, npv_sav)

        best_smr = max(savings, key=lambda x: savings[x][1])
        bg = COLOR["teal_light"] if alt else COLOR["white"]
        vals = [
            td["name"],
            td["type_link"].upper(),
            f"{td['annual_power_gwh']:,}",
            f"${td['grid_price_per_mwh']}",
            f"${grid_10y:,.0f}M",
            f"${savings['nuscale'][0]:.0f}/MWh",
            f"${savings['nuscale'][1]:,.0f}M",
            f"${savings['rolls_royce'][0]:.0f}/MWh",
            f"${savings['rolls_royce'][1]:,.0f}M",
            f"${savings['kairos'][0]:.0f}/MWh",
            f"${savings['kairos'][1]:,.0f}M",
            smr_types[best_smr]["name"].split()[0],
        ]
        row = _data_row(ws, row, vals, bg)
        alt = not alt

    row += 1

    # ─ [3] SMR 시나리오 전력비 절감 요약 ─
    row = _section_title(ws, row, 1,
        "📊 SMR Scenario Power Cost Summary", 8, COLOR["amber"])
    headers3 = [
        "Scenario", "SMR Type", "Penetration",
        "Add. CAPEX ($M)", "Opex Δ",
        "IRR Alpha (+pp)", "CO₂ Reduction", "Timeline"
    ]
    row = _table_header_row(ws, row, headers3, COLOR["amber"])

    scen_row_map = {}
    for sk, sd in scenarios.items():
        smr_name = "—" if not sd["smr_type"] else \
            smr_types.get(sd["smr_type"], {}).get("name", sd["smr_type"])
        bg = COLOR["amber_light"] if sk == "smr_partial" else COLOR["white"]
        vals = [
            sd["name"],
            smr_name,
            f"{sd['smr_penetration']*100:.0f}%",
            f"${sd['capex_addition_m']:,}M",
            f"{sd['opex_delta_pct']*100:+.0f}%",
            f"+{sd['irr_alpha_pp']:.1f}pp",
            f"{sd['co2_reduction_pct']}%",
            sd["timeline"],
        ]
        scen_row_map[sk] = row
        row = _data_row(ws, row, vals, bg)

    row += 1

    # ─ [4] 전력가 민감도 히트맵 (±$20/MWh) ─
    row = _section_title(ws, row, 1,
        "🔥 Electricity Price Sensitivity — NPV Savings Heatmap ($M)", 10,
        COLOR["red"])

    elec_sens = sens["electricity_price_sweep"]
    base_p = elec_sens["base_price_per_mwh"]
    r_range = elec_sens["range_per_mwh"]
    step = elec_sens["step_per_mwh"]
    price_steps = []
    p = base_p - r_range
    while p <= base_p + r_range + 0.01:
        price_steps.append(round(p, 1))
        p += step

    # 헤더: 전력가 × SMR 유형
    hmap_headers = ["Elec Price\n($/MWh)"] + [
        f"{smr_types[sk]['name'].split()[0]}\nNPV ($M)" for sk in smr_keys
    ]
    row = _table_header_row(ws, row, hmap_headers, COLOR["red"])

    # 대표 티어: AI DC US (대용량)
    ref_tier = dc_tiers["ai_datacenter_us"]
    hmap_data_start = row
    for p_val in price_steps:
        is_base = abs(p_val - base_p) < 0.01
        bg = COLOR["amber_light"] if is_base else COLOR["white"]
        row_vals = [f"${p_val:.0f}"]
        for sk, lcoe in smr_lcoe_map.items():
            npv_s = calc_npv_savings(
                p_val, lcoe, 0.8,
                ref_tier["annual_power_gwh"], dr, proj_y
            )
            row_vals.append(f"${npv_s:,.0f}M")
        row = _data_row(ws, row, row_vals, bg)

    row += 1

    # ─ [5] SMR CAPEX 충격 시나리오 ─
    row = _section_title(ws, row, 1,
        "⚡ SMR CAPEX Shock Scenarios (±20%)", 8, COLOR["gray_dark"])
    shock_headers = [
        "CAPEX Shock", "NuScale CAPEX ($B)", "NuScale LCOE",
        "RR CAPEX ($B)", "RR LCOE",
        "Kairos CAPEX ($B)", "Kairos LCOE",
        "IRR Alpha Impact"
    ]
    row = _table_header_row(ws, row, shock_headers, COLOR["gray_dark"])

    capex_shocks = sens["smr_capex_shock"]["shocks"]
    smr_base_full = scenarios["smr_full"]
    for shock in capex_shocks:
        is_base = shock == 0.0
        bg = COLOR["blue_light"] if is_base else COLOR["white"]
        row_vals = [f"{shock:+.0%}"]
        for sk in smr_keys:
            sm = smr_types[sk]
            shocked_capex = sm["capex_b_usd"] * (1 + shock)
            sm_shocked = {**sm, "capex_b_usd": shocked_capex}
            shocked_lcoe = calc_lcoe(sm_shocked)
            row_vals += [
                f"${shocked_capex:.2f}B",
                f"${shocked_lcoe:.0f}/MWh",
            ]
        # IRR Alpha 영향: CAPEX +20% → Alpha -0.5pp 선형 가정
        alpha_base = smr_base_full["irr_alpha_pp"]
        alpha_impact = alpha_base - shock * 2.5  # 선형 모델
        row_vals.append(f"+{alpha_impact:.1f}pp")
        row = _data_row(ws, row, row_vals, bg)

    row += 1

    # ─ NPV 절감액 Bar 차트 ─
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "SMR NPV Power Cost Savings by Facility Tier ($M, SMR Full 80%)"
    chart.y_axis.title = "NPV Savings ($M)"
    chart.style = 10
    chart.width = 28
    chart.height = 13

    n_tiers = len(dc_tiers)
    n_smrs = len(smr_keys)
    smr_colors = [COLOR["purple"], COLOR["teal"], COLOR["amber"]]

    for si, sk in enumerate(smr_keys):
        ref = Reference(
            ws,
            min_col=6 + si * 2,  # NuScale=6, RR=8, Kairos=10
            min_row=tier_smr_table_start,
            max_row=tier_smr_table_start + n_tiers - 1,
        )
        chart.add_data(ref, titles_from_data=False)
        chart.series[si].title.v = smr_types[sk]["name"].split()[0]
        chart.series[si].graphicalProperties.solidFill = smr_colors[si]

    cat_ref = Reference(
        ws, min_col=1,
        min_row=tier_smr_table_start,
        max_row=tier_smr_table_start + n_tiers - 1,
    )
    chart.set_categories(cat_ref)
    ws.add_chart(chart, f"A{row}")

    _col_widths(ws, {
        "A": 26, "B": 18, "C": 14, "D": 14, "E": 16,
        "F": 16, "G": 18, "H": 16, "I": 18,
        "J": 14, "K": 18, "L": 16, "M": 16, "N": 4,
    })

    print("  ✅ Sheet 8: SMR Power Cost Analysis")
    return smr_lcoe_map


# ──────────────────────────────────────────────────────────
# Sheet 9: Integrated DCF (Type A/B/C + SMR Alpha)
# ──────────────────────────────────────────────────────────
def build_sheet_integrated_dcf(wb, cfg: dict, smr_lcoe_map: dict):
    ws = wb.create_sheet(title="9. Integrated DCF")
    ws.sheet_view.showGridLines = False

    meta = cfg["meta"]
    scenarios = cfg["scenarios"]
    dc_tiers = cfg["data_center_tiers"]
    exaflops = cfg["exaflops_tco"]
    dr = meta["discount_rate"]
    proj_y = meta["projection_years"]

    # ─ 타이틀 ─
    ws.merge_cells("A1:N1")
    tc = ws["A1"]
    tc.value = "Integrated DCF — Type A/B/C + SMR Power Alpha | AstraChips LP Fund"
    tc.fill = _fill(COLOR["navy"])
    tc.font = _font(bold=True, size=15, color=COLOR["white"])
    tc.alignment = _align("center")
    _row_height(ws, 1, 36)

    ws.merge_cells("A2:N2")
    sub = ws["A2"]
    sub.value = (
        "Base IRR (Type A 13% / B 38% / C 42%) + SMR Alpha 반영 → "
        f"Integrated Portfolio MOIC | Discount Rate {dr*100:.0f}% | {proj_y}Y horizon"
    )
    sub.fill = _fill(COLOR["gray_dark"])
    sub.font = _font(size=10, color=COLOR["white"], italic=True)
    sub.alignment = _align("center")
    _row_height(ws, 2, 18)

    row = 4

    # 베이스 IRR 정의 (P2와 동일)
    BASE_IRR = {"type_a": 13.0, "type_b": 38.0, "type_c": 42.0}
    BASE_INVEST = {"type_a": 400, "type_b": 350, "type_c": 250}  # $M
    BASE_MOIC   = {"type_a": 1.11, "type_b": 8.90, "type_c": 11.50}
    TYPE_LABELS = {
        "type_a": "Glass 수직통합 (Korea)",
        "type_b": "HBM 패키징 레버리지 (Taiwan)",
        "type_c": "US/EU 분산 헤지",
    }
    TYPE_TIER = {"type_a": "hbm_fab", "type_b": "ai_datacenter_tw", "type_c": "ai_datacenter_us"}
    TYPE_COLORS = {
        "type_a": (COLOR["blue"], COLOR["blue_light"]),
        "type_b": (COLOR["teal"], COLOR["teal_light"]),
        "type_c": (COLOR["amber"], COLOR["amber_light"]),
    }

    # ─ [1] SMR 시나리오별 통합 IRR/MOIC 테이블 ─
    row = _section_title(ws, row, 1,
        "🔗 Integrated IRR & MOIC: Base + SMR Power Alpha", 12,
        COLOR["purple"])
    headers1 = [
        "Type", "Strategy",
        "Base IRR (%)", "Invest ($M)", "Base MOIC",
        "Grid Only\nIRR / MOIC",
        "SMR Partial\nIRR (+1.2pp)",
        "SMR Partial\nMOIC",
        "SMR Full\nIRR (+2.8pp)",
        "SMR Full\nMOIC",
        "Max Delta MOIC", "Rec. SMR"
    ]
    row = _table_header_row(ws, row, headers1, COLOR["purple"])

    integrated_table_start = row
    type_integrated = {}

    for tk in ["type_a", "type_b", "type_c"]:
        bg_main, bg_light = TYPE_COLORS[tk]
        base_irr = BASE_IRR[tk]
        invest = BASE_INVEST[tk]
        base_moic = BASE_MOIC[tk]

        smr_partial_irr = base_irr + scenarios["smr_partial"]["irr_alpha_pp"]
        smr_full_irr    = base_irr + scenarios["smr_full"]["irr_alpha_pp"]

        # MOIC 재계산: MOIC ≈ (1 + IRR/100)^horizon / (1 + base_irr/100)^horizon * base_moic
        def irr_to_moic_delta(irr_new, irr_base, base_moic, years):
            ratio = ((1 + irr_new/100) / (1 + irr_base/100)) ** years
            return round(base_moic * ratio, 2)

        moic_partial = irr_to_moic_delta(smr_partial_irr, base_irr, base_moic, proj_y)
        moic_full    = irr_to_moic_delta(smr_full_irr, base_irr, base_moic, proj_y)
        delta_moic   = round(moic_full - base_moic, 2)

        type_integrated[tk] = {
            "base_irr": base_irr,
            "smr_partial_irr": round(smr_partial_irr, 1),
            "smr_full_irr": round(smr_full_irr, 1),
            "base_moic": base_moic,
            "moic_partial": moic_partial,
            "moic_full": moic_full,
            "delta_moic": delta_moic,
            "invest": invest,
        }

        vals = [
            tk.upper().replace("_", " "),
            TYPE_LABELS[tk],
            f"{base_irr:.1f}%",
            f"${invest:,}M",
            f"{base_moic:.2f}x",
            f"{base_irr:.1f}% / {base_moic:.2f}x",
            f"{smr_partial_irr:.1f}%",
            f"{moic_partial:.2f}x",
            f"{smr_full_irr:.1f}%",
            f"{moic_full:.2f}x",
            f"+{delta_moic:.2f}x",
            "NuScale" if tk == "type_c" else "Kairos",
        ]
        row = _data_row(ws, row, vals, bg_light)

    row += 1

    # ─ [2] 포트폴리오 통합 MOIC (Balanced 배분 40/35/25) ─
    row = _section_title(ws, row, 1,
        "📈 Portfolio Integrated MOIC — Balanced Allocation (40/35/25)", 10,
        COLOR["teal"])
    weights = {"type_a": 0.40, "type_b": 0.35, "type_c": 0.25}

    port_headers = [
        "SMR Scenario", "Type A MOIC", "Type B MOIC", "Type C MOIC",
        "Portfolio MOIC", "ΔMOIC vs Grid", "Portfolio Return ($M)",
        "SMR Add. CAPEX ($M)", "Net Alpha ($M)", "Rec."
    ]
    row = _table_header_row(ws, row, port_headers, COLOR["teal"])

    port_scenarios = [
        ("Grid Only",   "base_moic",    0,   COLOR["white"]),
        ("SMR Partial", "moic_partial", 360, COLOR["teal_light"]),
        ("SMR Full",    "moic_full",   930, COLOR["green_light"]),
    ]
    total_invest = sum(BASE_INVEST.values())  # $1,000M

    port_table_start = row
    base_port_moic = None
    for scen_label, moic_key, add_capex, bg in port_scenarios:
        port_moic = sum(
            weights[tk] * type_integrated[tk][moic_key] for tk in weights
        )
        port_moic = round(port_moic, 2)
        if base_port_moic is None:
            base_port_moic = port_moic
        delta = round(port_moic - base_port_moic, 2)
        port_return = round(port_moic * total_invest, 0)
        net_alpha = round((port_return - base_port_moic * total_invest) - add_capex, 0)
        is_rec = scen_label == "SMR Partial"
        vals = [
            scen_label,
            f"{type_integrated['type_a'][moic_key]:.2f}x",
            f"{type_integrated['type_b'][moic_key]:.2f}x",
            f"{type_integrated['type_c'][moic_key]:.2f}x",
            f"{port_moic:.2f}x",
            f"+{delta:.2f}x" if delta >= 0 else f"{delta:.2f}x",
            f"${port_return:,.0f}M",
            f"${add_capex:,}M" if add_capex else "—",
            f"${net_alpha:,.0f}M" if add_capex else "—",
            "⭐ REC" if is_rec else "",
        ]
        row = _data_row(ws, row, vals, bg)

    row += 1

    # ─ [3] ExaFLOPS TCO 단위 분석 ─
    row = _section_title(ws, row, 1,
        "🖥️  ExaFLOPS TCO Analysis — Power Cost Impact", 8,
        COLOR["navy"])
    exa_headers = [
        "Metric", "Unit",
        "Grid Only", "SMR Partial", "SMR Full",
        "Target", "SMR Full vs Target", "Notes"
    ]
    row = _table_header_row(ws, row, exa_headers, COLOR["navy"])

    exa = exaflops
    # 대표 AI DC: US (Type C)
    ref_tier = dc_tiers["ai_datacenter_us"]
    it_mw = ref_tier["it_load_mw"]
    exaflops_val = it_mw * exa["exaflops_per_mw_it"]
    hbm_tb = exaflops_val * exa["hbm_tb_per_exaflops"]
    grid_price = ref_tier["grid_price_per_mwh"]
    annual_gwh = ref_tier["annual_power_gwh"]

    grid_annual_cost_m = grid_price * annual_gwh * 1e3 / 1e6
    partial_annual_cost_m = grid_annual_cost_m * (1 - scenarios["smr_partial"]["opex_delta_pct"] * (-1))
    full_annual_cost_m = grid_annual_cost_m * (1 + scenarios["smr_full"]["opex_delta_pct"])

    tco_target_m = exa["target_tco_per_exaflops_m"] * exaflops_val
    pwr_share = exa["power_cost_share_pct"]

    exa_rows = [
        ("ExaFLOPS Capacity",    "ExaFLOPS",
         f"{exaflops_val:.1f}", f"{exaflops_val:.1f}", f"{exaflops_val:.1f}",
         "—", "—", f"{it_mw}MW IT Load"),
        ("HBM Required",         "TB",
         f"{hbm_tb:.0f}", f"{hbm_tb:.0f}", f"{hbm_tb:.0f}",
         "—", "—", f"{exa['hbm_tb_per_exaflops']} TB/ExaFLOPS"),
        ("Annual Power Cost",     "$M/yr",
         f"${grid_annual_cost_m:.1f}M",
         f"${partial_annual_cost_m:.1f}M",
         f"${full_annual_cost_m:.1f}M",
         "—", "—", "Grid escalation excl."),
        ("TCO (10Y)",             "$M",
         f"${grid_annual_cost_m*proj_y:.0f}M",
         f"${partial_annual_cost_m*proj_y:.0f}M",
         f"${full_annual_cost_m*proj_y:.0f}M",
         f"${tco_target_m:.0f}M",
         f"{(full_annual_cost_m*proj_y/tco_target_m - 1)*100:+.1f}%",
         f"{pwr_share*100:.0f}% power share of TCO"),
        ("TCO/ExaFLOPS",         "$M/ExaFLOPS",
         f"${grid_annual_cost_m*proj_y/exaflops_val:.0f}M",
         f"${partial_annual_cost_m*proj_y/exaflops_val:.0f}M",
         f"${full_annual_cost_m*proj_y/exaflops_val:.0f}M",
         f"${exa['target_tco_per_exaflops_m']:.0f}M",
         "—", "Power component only"),
        ("SMR TCO Reduction",    "%",
         "0%",
         f"{exa['smr_tco_reduction_pct']['smr_partial']:.1f}%",
         f"{exa['smr_tco_reduction_pct']['smr_full']:.1f}%",
         "—", "—", "vs Grid Only baseline"),
    ]
    alt = False
    for er in exa_rows:
        bg = COLOR["gray_light"] if alt else COLOR["white"]
        row = _data_row(ws, row, list(er), bg)
        alt = not alt

    row += 1

    # ─ Portfolio MOIC 비교 라인 차트 ─
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Portfolio MOIC: Grid Only vs SMR Partial vs SMR Full"
    chart2.y_axis.title = "Portfolio MOIC (x)"
    chart2.style = 10
    chart2.width = 22
    chart2.height = 12

    for col_offset, series_label in [
        (5, "Type A MOIC"),
        (6, "Type B MOIC"),
        (7, "Type C MOIC"),
        (8, "Portfolio MOIC"),
    ]:
        ref = Reference(ws, min_col=col_offset, max_col=col_offset,
                        min_row=port_table_start, max_row=port_table_start + 2)
        chart2.add_data(ref, titles_from_data=False)
        chart2.series[-1].title.v = series_label

    cat_ref = Reference(ws, min_col=1,
                        min_row=port_table_start, max_row=port_table_start + 2)
    chart2.set_categories(cat_ref)
    ws.add_chart(chart2, f"A{row}")

    _col_widths(ws, {
        "A": 24, "B": 32, "C": 14, "D": 14, "E": 12,
        "F": 20, "G": 18, "H": 16, "I": 18, "J": 16,
        "K": 16, "L": 14, "M": 14, "N": 4,
    })

    print("  ✅ Sheet 9: Integrated DCF (Type A/B/C + SMR Alpha)")


# ──────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="SMR Power Model v1.0 — Sheet 8~9 Generator"
    )
    parser.add_argument(
        "--input",
        default="reports/AstraChips_Finance_Model.xlsx",
        help="기존 Excel 모델 경로 (Sheet 1~7 포함)"
    )
    parser.add_argument(
        "--config",
        default="plugins/finance/smr_power_config.yaml",
        help="SMR 파라미터 YAML 경로"
    )
    parser.add_argument(
        "--output",
        default="reports/AstraChips_Finance_Model_v1.2.xlsx",
        help="출력 XLSX 경로 (Sheet 1~9)"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="계산만 수행, 파일 저장 안 함")
    args = parser.parse_args()

    # YAML 로드
    cfg_path = Path(args.config)
    if not cfg_path.exists():
        print(f"[ERROR] Config not found: {cfg_path}")
        sys.exit(1)
    with open(cfg_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    print(f"\n📂 SMR config loaded: {cfg_path}")
    print(f"   SMR types: {list(cfg['smr_types'].keys())}")
    print(f"   DC tiers:  {list(cfg['data_center_tiers'].keys())}\n")

    # 기존 Excel 로드
    in_path = Path(args.input)
    if not in_path.exists():
        print(f"[WARN] Base Excel not found: {in_path}")
        print("  → 새 워크북 생성 (Sheet 8~9만 포함)")
        import openpyxl
        wb = openpyxl.Workbook()
        # 빈 Sheet 1 제거 후 다시 사용하지 않음
        default_ws = wb.active
        default_ws.title = "_placeholder"
    else:
        import openpyxl
        wb = openpyxl.load_workbook(in_path)
        print(f"📊 Base model loaded: {in_path}")
        print(f"   Existing sheets: {[s.title for s in wb.worksheets]}")

    print("\n🔧 Building SMR Power Model sheets...")

    # Sheet 8
    smr_lcoe_map = build_sheet_smr_power(wb, cfg)

    # Sheet 9
    build_sheet_integrated_dcf(wb, cfg, smr_lcoe_map)

    if args.dry_run:
        print("\n✅ Dry-run complete — 파일 저장 생략")
        print(f"   LCOE Map: {smr_lcoe_map}")
        return 0

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"\n✅ Excel model v1.2 saved: {out_path}")
    print(f"   Total sheets: {[s.title for s in wb.worksheets]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
